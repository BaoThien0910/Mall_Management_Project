# File: app/services/excel_import_service.py
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.constants.billing import (
    IMPORT_ALLOWED_FINANCIAL_TYPES,
    IMPORT_FORBIDDEN_METER_TYPES,
)
from app.exceptions.business_exceptions import BadRequestException
from app.models.dulieu_import_taichinh import DuLieuImportTaiChinh
from app.models.hopdong import HopDong
from app.services._common import generate_code, get_column, get_value

EXPECTED_HEADERS = ["MAHD", "THANG", "NAM", "LOAIKHOAN", "SOTIEN", "GHICHU"]


def _current_employee_id(current_user: Any) -> str:
    ma_nv = get_value(current_user, ["ma_nv", "ma_nhan_vien", "manv"])
    if not ma_nv:
        raise BadRequestException("Tài khoản hiện tại không gắn với nhân viên")
    return ma_nv


def _normalize_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _to_decimal(value: Any) -> Decimal:
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError):
        raise ValueError("Số tiền không hợp lệ")


def _validate_headers(headers: List[str]) -> None:
    normalized_headers = [_normalize_text(header).upper() for header in headers]
    if normalized_headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise BadRequestException(
            "File Excel không đúng cấu trúc. Header bắt buộc: "
            "MAHD, THANG, NAM, LOAIKHOAN, SOTIEN, GHICHU"
        )


def _build_import_row(
    ma_hd: str,
    thang: int,
    nam: int,
    loai_khoan: str,
    so_tien: Decimal,
    ghi_chu: Optional[str],
    ma_nhan_vien_import: str,
    ten_file: Optional[str],
    dong_excel: int,
    trang_thai: str,
    loi_chi_tiet: Optional[str],
) -> DuLieuImportTaiChinh:
    return DuLieuImportTaiChinh(
        ma_import=generate_code("IMP"),
        ma_hop_dong=ma_hd,
        thang=thang,
        nam=nam,
        loai_khoan=loai_khoan,
        so_tien=so_tien,
        ghi_chu=ghi_chu,
        ma_nhan_vien_import=ma_nhan_vien_import,
        ten_file=ten_file,
        dong_excel=dong_excel,
        trang_thai=trang_thai,
        loi_chi_tiet=loi_chi_tiet,
    )


def import_financial_excel(
    db: Session,
    file: UploadFile,
    current_user: Any,
) -> Dict[str, Any]:
    """Import dữ liệu tài chính từ Excel.

    Excel chỉ nhận: Tiền thuê, Phí bảo trì, Hoàn trả.
    Tiền điện/nước được tính từ CHISODIENNUOC, không nhập qua Excel.
    """
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise BadRequestException("Chỉ hỗ trợ file Excel định dạng .xlsx")

    ma_nhan_vien_import = _current_employee_id(current_user)

    try:
        file.file.seek(0)
        file_contents = file.file.read()
        workbook = load_workbook(BytesIO(file_contents), data_only=True)
    except Exception:
        raise BadRequestException("Không đọc được file Excel")

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise BadRequestException("File Excel không có dữ liệu")

    headers = list(rows[0])
    _validate_headers(headers)

    hopdong_id_col = get_column(HopDong, ["ma_hop_dong", "ma_hd"])

    # 1. Thu thập tất cả các mã hợp đồng từ file Excel để truy vấn một lần
    excel_ma_hds = set()
    for row in rows[1:]:
        if row is not None and not all(value is None for value in row):
            values = list(row)
            if len(values) > 0 and values[0]:
                excel_ma_hds.add(_normalize_text(values[0]))

    # 2. Lấy danh sách các khoản đã import thành công hoặc đang dùng của các hợp đồng này
    existing_keys = set()
    if excel_ma_hds:
        ma_hd_col = get_column(DuLieuImportTaiChinh, ["ma_hop_dong", "ma_hd"])
        stmt = select(DuLieuImportTaiChinh).where(
            ma_hd_col.in_(list(excel_ma_hds)),
            DuLieuImportTaiChinh.trang_thai.in_(["Hợp lệ", "Đã dùng tính công nợ"])
        )
        existing_records = db.execute(stmt).scalars().all()
        for r in existing_records:
            existing_keys.add((r.ma_hop_dong, r.thang, r.nam, r.loai_khoan))

    seen_in_file = set()

    total_rows = 0
    valid_rows = 0
    invalid_rows = 0
    error_details: List[Dict[str, Any]] = []
    import_items: List[DuLieuImportTaiChinh] = []

    for excel_row_number, row in enumerate(rows[1:], start=2):
        if row is None or all(value is None for value in row):
            continue

        total_rows += 1
        values = list(row) + [None] * (len(EXPECTED_HEADERS) - len(row))
        ma_hd = _normalize_text(values[0])
        raw_thang = values[1]
        raw_nam = values[2]
        loai_khoan = _normalize_text(values[3])
        raw_so_tien = values[4]
        ghi_chu = _normalize_text(values[5]) or None

        row_errors: List[str] = []

        if not ma_hd:
            row_errors.append("Thiếu mã hợp đồng")

        try:
            thang = int(raw_thang)
        except (TypeError, ValueError):
            thang = 0
            row_errors.append("Tháng không hợp lệ")

        try:
            nam = int(raw_nam)
        except (TypeError, ValueError):
            nam = 0
            row_errors.append("Năm không hợp lệ")

        if thang < 1 or thang > 12:
            row_errors.append("Tháng phải nằm trong khoảng 1 đến 12")
        if nam < 2000 or nam > 2100:
            row_errors.append("Năm phải nằm trong khoảng 2000 đến 2100")

        try:
            so_tien = _to_decimal(raw_so_tien)
            if so_tien <= 0:
                row_errors.append("Số tiền phải lớn hơn 0")
        except ValueError as exc:
            so_tien = Decimal("0")
            row_errors.append(str(exc))

        if loai_khoan in IMPORT_FORBIDDEN_METER_TYPES:
            row_errors.append(
                "Tiền điện và tiền nước được tính từ chức năng chỉ số điện nước, không nhập qua Excel"
            )
        elif loai_khoan not in IMPORT_ALLOWED_FINANCIAL_TYPES:
            row_errors.append(
                "Loại khoản không hợp lệ. Chỉ cho phép: Tiền thuê, Phí bảo trì, Hoàn trả"
            )

        hop_dong = None
        if ma_hd:
            hop_dong = db.execute(
                select(HopDong).where(hopdong_id_col == ma_hd)
            ).scalars().first()
            if hop_dong is None:
                row_errors.append("Mã hợp đồng không tồn tại")

        # 3. Check trùng lặp dữ liệu (trong file và trong hệ thống)
        if ma_hd and (1 <= thang <= 12) and (2000 <= nam <= 2100) and loai_khoan:
            row_key = (ma_hd, thang, nam, loai_khoan)
            if row_key in seen_in_file:
                row_errors.append("Dòng bị trùng lặp trong file Excel")
            else:
                seen_in_file.add(row_key)

            if row_key in existing_keys:
                row_errors.append(f"Khoản thu cho hợp đồng này đã tồn tại trong hệ thống ở kỳ {thang}/{nam}")

        if row_errors:
            invalid_rows += 1
            error_details.append(
                {
                    "dong_excel": excel_row_number,
                    "ma_hop_dong": ma_hd,
                    "loi": "; ".join(row_errors),
                }
            )
            # Chỉ lưu dòng lỗi nếu thỏa mãn các check constraint của database (ví dụ: tháng 1-12, loại khoản hợp lệ, số tiền > 0, v.v.).
            db_thang_valid = 1 <= thang <= 12
            db_loai_khoan_valid = loai_khoan in {
                "Tiền thuê",
                "Tiền điện",
                "Tiền nước",
                "Phí bảo trì",
                "Hoàn trả",
            }
            db_so_tien_valid = so_tien > 0

            if (
                hop_dong is not None
                and db_thang_valid
                and nam
                and db_loai_khoan_valid
                and db_so_tien_valid
            ):
                import_items.append(
                    _build_import_row(
                        ma_hd=ma_hd,
                        thang=thang,
                        nam=nam,
                        loai_khoan=loai_khoan,
                        so_tien=so_tien,
                        ghi_chu=ghi_chu,
                        ma_nhan_vien_import=ma_nhan_vien_import,
                        ten_file=filename,
                        dong_excel=excel_row_number,
                        trang_thai="Lỗi",
                        loi_chi_tiet="; ".join(row_errors)[:255],
                    )
                )
            continue

        valid_rows += 1
        import_items.append(
            _build_import_row(
                ma_hd=ma_hd,
                thang=thang,
                nam=nam,
                loai_khoan=loai_khoan,
                so_tien=so_tien,
                ghi_chu=ghi_chu,
                ma_nhan_vien_import=ma_nhan_vien_import,
                ten_file=filename,
                dong_excel=excel_row_number,
                trang_thai="Hợp lệ",
                loi_chi_tiet=None,
            )
        )

    try:
        db.add_all(import_items)
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "ten_file": filename,
        "tong_so_dong": total_rows,
        "so_dong_hop_le": valid_rows,
        "so_dong_loi": invalid_rows,
        "chi_tiet_loi": error_details,
    }


def _serialize_import_item(item: DuLieuImportTaiChinh) -> Dict[str, Any]:
    return {
        "ma_import": get_value(item, ["ma_import"], ""),
        "ma_hop_dong": get_value(item, ["ma_hop_dong", "ma_hd"], ""),
        "thang": item.thang,
        "nam": item.nam,
        "loai_khoan": get_value(item, ["loai_khoan"], ""),
        "so_tien": get_value(item, ["so_tien"], Decimal("0")),
        "ghi_chu": get_value(item, ["ghi_chu"], None),
        "ten_file": get_value(item, ["ten_file"], None),
        "dong_excel": get_value(item, ["dong_excel"], None),
        "trang_thai": get_value(item, ["trang_thai"], ""),
        "loi_chi_tiet": get_value(item, ["loi_chi_tiet"], None),
        "thoi_gian_import": get_value(item, ["thoi_gian_import"], None),
    }


def list_import_history(
    db: Session,
    filters: Optional[Any] = None,
    current_user: Optional[Any] = None,
) -> Dict[str, Any]:
    stmt = select(DuLieuImportTaiChinh)

    ma_hop_dong = get_value(filters, ["ma_hop_dong", "ma_hd"], None) if filters else None
    keyword = get_value(filters, ["keyword"], None) if filters else None
    loai_khoan = get_value(filters, ["loai_khoan"], None) if filters else None
    thang = get_value(filters, ["thang"], None) if filters else None
    nam = get_value(filters, ["nam"], None) if filters else None
    trang_thai = get_value(filters, ["trang_thai"], None) if filters else None
    page = get_value(filters, ["page"], 1) if filters else 1
    page_size = get_value(filters, ["page_size"], 10) if filters else 10

    if ma_hop_dong:
        stmt = stmt.where(get_column(DuLieuImportTaiChinh, ["ma_hop_dong", "ma_hd"]).ilike(f"%{ma_hop_dong}%"))
    if keyword:
        stmt = stmt.where(get_column(DuLieuImportTaiChinh, ["ma_hop_dong", "ma_hd"]).ilike(f"%{keyword}%"))
    if loai_khoan:
        stmt = stmt.where(DuLieuImportTaiChinh.loai_khoan == (loai_khoan.value if hasattr(loai_khoan, "value") else loai_khoan))
    if thang:
        stmt = stmt.where(DuLieuImportTaiChinh.thang == thang)
    if nam:
        stmt = stmt.where(DuLieuImportTaiChinh.nam == nam)
    if trang_thai:
        stmt = stmt.where(DuLieuImportTaiChinh.trang_thai == (trang_thai.value if hasattr(trang_thai, "value") else trang_thai))

    total = db.execute(select(func.count()).select_from(stmt.subquery())).scalar_one()
    items = db.execute(
        stmt.order_by(DuLieuImportTaiChinh.thoi_gian_import.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).scalars().all()

    return {
        "items": [_serialize_import_item(item) for item in items],
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": (total + page_size - 1) // page_size if total else 0,
        },
    }


def batch_delete_imports(
    db: Session,
    ids: List[str],
    current_user: Any,
) -> Dict[str, Any]:
    """Xóa hàng loạt các bản ghi import tài chính.
    Chỉ cho phép xóa các bản ghi ở trạng thái 'Hợp lệ' hoặc 'Lỗi'.
    Không cho phép xóa bản ghi ở trạng thái 'Đã dùng tính công nợ' để bảo toàn dữ liệu.
    """
    if not ids:
        raise BadRequestException("Không có bản ghi nào được chọn để xóa")

    ma_import_col = get_column(DuLieuImportTaiChinh, ["ma_import", "maimport"])
    stmt = select(DuLieuImportTaiChinh).where(ma_import_col.in_(ids))
    records = db.execute(stmt).scalars().all()

    # Kiểm tra tồn tại
    if len(records) != len(ids):
        found_ids = {get_value(r, ["ma_import", "maimport"]) for r in records}
        missing_ids = set(ids) - found_ids
        raise BadRequestException(f"Một số bản ghi không tồn tại: {list(missing_ids)}")

    # Kiểm tra trạng thái đã sử dụng
    for r in records:
        trang_thai = get_value(r, ["trang_thai"])
        if trang_thai == "Đã dùng tính công nợ":
            ma_imp = get_value(r, ["ma_import", "maimport"])
            raise BadRequestException(
                f"Không thể xóa bản ghi {ma_imp} vì đã được sử dụng để tính công nợ"
            )

    # Xóa
    deleted_count = 0
    try:
        for r in records:
            db.delete(r)
            deleted_count += 1
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "so_luong_xoa": deleted_count,
    }


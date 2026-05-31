# File: app/services/financial_report_service.py
import uuid
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, select
from sqlalchemy.orm import Session, selectinload

from app.constants.statuses import AuditAction, HoaDonStatus
from app.exceptions.business_exceptions import BadRequestException, ConflictException, NotFoundException
from app.models import BaoCaoTaiChinh, BaoCaoTaiChinhChiTiet, CongNo, HoaDon, HopDong
from app.schemas.baocaotaichinh import BaoCaoTaiChinhCreate, BaoCaoTaiChinhFilter
from app.services.audit_service import write_audit_log
from app.utils.datetime_helper import current_datetime
from app.utils.pagination import calculate_offset, calculate_total_pages, normalize_pagination
from app.utils.transaction import transaction_context


MONEY_HEADERS = {
    "tien_thue",
    "tien_dien",
    "tien_nuoc",
    "tien_hoan_tra",
    "chi_phi_bao_tri",
    "tong_tien",
    "da_thanh_toan",
    "no",
}


def _generate_code(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12].upper()}"


def _user_attr(current_user: Any, short_name: str, long_name: str) -> Any:
    value = getattr(current_user, short_name, None)
    return value if value is not None else getattr(current_user, long_name, None)


def _ky_chot(thang: int, nam: int) -> str:
    return f"{thang:02d}/{nam}"


def _to_decimal(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _money(value: Any) -> Decimal:
    return _to_decimal(value).quantize(Decimal("0.01"))


def _get_report_or_404(db: Session, ma_bc: str) -> BaoCaoTaiChinh:
    report = db.execute(
        select(BaoCaoTaiChinh)
        .options(selectinload(BaoCaoTaiChinh.chi_tiets))
        .where(BaoCaoTaiChinh.ma_bao_cao == ma_bc)
    ).scalars().first()

    if report is None:
        raise NotFoundException("Không tìm thấy báo cáo tài chính")

    return report


def _build_summary(details: List[BaoCaoTaiChinhChiTiet]) -> Dict[str, Any]:
    tong = sum((_money(item.tong_tien) for item in details), Decimal("0"))
    tong_tt = sum((_money(item.da_thanh_toan) for item in details), Decimal("0"))
    tong_no = sum((_money(item.no) for item in details), Decimal("0"))

    return {
        "tong_so_hd": len(details),
        "tong_so_hd_con_no": sum(1 for item in details if _money(item.no) > 0),
        "tong": _money(tong),
        "tong_tt": _money(tong_tt),
        "tong_no": _money(tong_no),
    }


def _detail_response(report: BaoCaoTaiChinh) -> Dict[str, Any]:
    return {
        "bao_cao": report,
        "chi_tiet": report.chi_tiets,
        "thong_ke": _build_summary(report.chi_tiets),
    }


def _successful_payment_map(db: Session, cong_no_ids: List[str]) -> Dict[str, Decimal]:
    if not cong_no_ids:
        return {}

    rows = db.execute(
        select(HoaDon.ma_cong_no, func.coalesce(func.sum(HoaDon.so_tien), 0))
        .where(
            HoaDon.ma_cong_no.in_(cong_no_ids),
            HoaDon.trang_thai == HoaDonStatus.THANH_CONG.value,
        )
        .group_by(HoaDon.ma_cong_no)
    ).all()

    return {ma_cong_no: _money(total) for ma_cong_no, total in rows}


def create_financial_report(
    db: Session,
    payload: BaoCaoTaiChinhCreate,
    current_user: Any,
) -> Dict[str, Any]:
    """Lập báo cáo tài chính theo tháng/năm, tự tổng hợp từ CONGNO + HOPDONG + HOADON."""
    ma_nhan_vien_lap = _user_attr(current_user, "ma_nv", "ma_nhan_vien")
    if not ma_nhan_vien_lap:
        raise BadRequestException("Không xác định được nhân viên lập báo cáo")

    existed = db.execute(
        select(BaoCaoTaiChinh).where(
            BaoCaoTaiChinh.thang == payload.thang,
            BaoCaoTaiChinh.nam == payload.nam,
        )
    ).scalars().first()
    if existed is not None:
        raise ConflictException(f"Báo cáo tài chính kỳ {_ky_chot(payload.thang, payload.nam)} đã tồn tại")

    rows = db.execute(
        select(CongNo, HopDong)
        .join(HopDong, HopDong.ma_hop_dong == CongNo.ma_hop_dong)
        .where(CongNo.thang == payload.thang, CongNo.nam == payload.nam)
        .order_by(CongNo.ma_hop_dong.asc())
    ).all()

    if not rows:
        raise BadRequestException(f"Không có dữ liệu công nợ cho kỳ {_ky_chot(payload.thang, payload.nam)}")

    payment_by_debt = _successful_payment_map(db, [cong_no.ma_cong_no for cong_no, _ in rows])
    ky = _ky_chot(payload.thang, payload.nam)

    report = BaoCaoTaiChinh(
        ma_bao_cao=_generate_code("BCTC"),
        ma_nhan_vien_lap=ma_nhan_vien_lap,
        thang=payload.thang,
        nam=payload.nam,
        ky_chot=ky,
        ngay_lap=current_datetime(),
    )

    for index, (cong_no, hop_dong) in enumerate(rows, start=1):
        da_thanh_toan = payment_by_debt.get(cong_no.ma_cong_no, Decimal("0"))
        tong_tien = _money(cong_no.tong_tien)
        no = max(tong_tien - da_thanh_toan, Decimal("0"))

        report.chi_tiets.append(
            BaoCaoTaiChinhChiTiet(
                ma_bao_cao=report.ma_bao_cao,
                stt=index,
                ma_hop_dong=cong_no.ma_hop_dong,
                ma_khach_thue=hop_dong.ma_khach_thue,
                ma_mat_bang=hop_dong.ma_mat_bang,
                ky=ky,
                tien_thue=_money(cong_no.tien_thue),
                tien_dien=_money(cong_no.tien_dien),
                tien_nuoc=_money(cong_no.tien_nuoc),
                tien_hoan_tra=_money(cong_no.tien_hoan),
                chi_phi_bao_tri=_money(cong_no.phi_bao_tri),
                tong_tien=tong_tien,
                da_thanh_toan=da_thanh_toan,
                no=no,
            )
        )

    actor_ma_tk = _user_attr(current_user, "ma_tk", "ma_tai_khoan")
    with transaction_context(db):
        db.add(report)
        db.flush()
        write_audit_log(
            db=db,
            ma_tk=actor_ma_tk,
            hanh_dong=AuditAction.TAO_MOI,
            doi_tuong="BAOCAOTAICHINH",
            ma_doi_tuong=report.ma_bao_cao,
            chi_tiet=f"Lập báo cáo tài chính kỳ {ky}",
        )

    report = _get_report_or_404(db, report.ma_bao_cao)
    return _detail_response(report)


def list_financial_reports(
    db: Session,
    filters: BaoCaoTaiChinhFilter,
    current_user: Optional[Any] = None,
) -> Dict[str, Any]:
    page, page_size = normalize_pagination(filters.page, filters.page_size)
    conditions: List[Any] = []

    if filters.thang is not None:
        conditions.append(BaoCaoTaiChinh.thang == filters.thang)
    if filters.nam is not None:
        conditions.append(BaoCaoTaiChinh.nam == filters.nam)

    stmt = select(BaoCaoTaiChinh)
    count_stmt = select(func.count()).select_from(BaoCaoTaiChinh)

    if conditions:
        clause = and_(*conditions)
        stmt = stmt.where(clause)
        count_stmt = count_stmt.where(clause)

    total = db.execute(count_stmt).scalar_one()
    items = db.execute(
        stmt.order_by(BaoCaoTaiChinh.nam.desc(), BaoCaoTaiChinh.thang.desc(), BaoCaoTaiChinh.ngay_lap.desc())
        .offset(calculate_offset(page, page_size))
        .limit(page_size)
    ).scalars().all()

    return {
        "items": items,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": calculate_total_pages(total, page_size),
        },
    }


def get_financial_report_detail(
    db: Session,
    ma_bc: str,
    current_user: Optional[Any] = None,
) -> Dict[str, Any]:
    report = _get_report_or_404(db, ma_bc)
    return _detail_response(report)


def export_financial_report_excel(
    db: Session,
    ma_bc: str,
    current_user: Optional[Any] = None,
) -> BytesIO:
    report = _get_report_or_404(db, ma_bc)
    summary = _build_summary(report.chi_tiets)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao tai chinh"

    headers = [
        "MAHD",
        "MAKH",
        "MAMB",
        "KY",
        "TIEN THUE",
        "TIEN DIEN",
        "TIEN NUOC",
        "TIEN HOAN TRA",
        "CHI PHI BAO TRI",
        "TONG TIEN",
        "DA THANH TOAN",
        "NO",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1).value = "BÁO CÁO TÀI CHÍNH"
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(2, 1).value = f"KỲ CHỐT: {report.ky_chot} | MABC: {report.ma_bao_cao} | MANV: {report.ma_nhan_vien_lap}"
    ws.cell(2, 1).alignment = Alignment(horizontal="center")

    header_row = 4
    fill = PatternFill("solid", fgColor="D9EAF7")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = fill

    for row_index, item in enumerate(report.chi_tiets, start=header_row + 1):
        values = [
            item.ma_hop_dong,
            item.ma_khach_thue,
            item.ma_mat_bang,
            item.ky,
            item.tien_thue,
            item.tien_dien,
            item.tien_nuoc,
            item.tien_hoan_tra,
            item.chi_phi_bao_tri,
            item.tong_tien,
            item.da_thanh_toan,
            item.no,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_index, col)
            cell.value = float(value) if isinstance(value, Decimal) else value
            if col >= 5:
                cell.number_format = '#,##0'

    stat_start = header_row + len(report.chi_tiets) + 3
    ws.cell(stat_start, 1).value = "THONG KE"
    ws.cell(stat_start, 1).font = Font(bold=True)

    stats = [
        ("TONG SO HD", summary["tong_so_hd"]),
        ("TONG SO HD CON NO", summary["tong_so_hd_con_no"]),
        ("TONG", summary["tong"]),
        ("TONG TT", summary["tong_tt"]),
        ("TONG NO", summary["tong_no"]),
    ]
    for offset, (label, value) in enumerate(stats, start=1):
        ws.cell(stat_start + offset, 1).value = label
        ws.cell(stat_start + offset, 2).value = float(value) if isinstance(value, Decimal) else value
        if isinstance(value, Decimal):
            ws.cell(stat_start + offset, 2).number_format = '#,##0'

    widths = [16, 14, 14, 10, 15, 15, 15, 18, 18, 15, 18, 15]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A5"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

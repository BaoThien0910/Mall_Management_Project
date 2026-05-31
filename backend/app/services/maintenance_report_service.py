# File: app/services/maintenance_report_service.py
import uuid
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, select
from sqlalchemy.orm import Session, selectinload

from app.constants.statuses import AuditAction, SuCoBaoTriStatus
from app.exceptions.business_exceptions import BadRequestException, ConflictException, NotFoundException
from app.models import BaoCaoBaoTri, BaoCaoBaoTriChiTiet, SuCoBaoTri
from app.schemas.baocaobaotri import BaoCaoBaoTriCreate, BaoCaoBaoTriFilter
from app.services.audit_service import write_audit_log
from app.utils.datetime_helper import current_datetime
from app.utils.pagination import calculate_offset, calculate_total_pages, normalize_pagination
from app.utils.transaction import transaction_context


def _generate_code(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12].upper()}"


def _user_attr(current_user: Any, short_name: str, long_name: str) -> Any:
    value = getattr(current_user, short_name, None)
    return value if value is not None else getattr(current_user, long_name, None)


def _ky_chot(thang: int, nam: int) -> str:
    return f"{thang:02d}/{nam}"


def _get_report_or_404(db: Session, ma_bc: str) -> BaoCaoBaoTri:
    report = db.execute(
        select(BaoCaoBaoTri)
        .options(selectinload(BaoCaoBaoTri.chi_tiets))
        .where(BaoCaoBaoTri.ma_bao_cao == ma_bc)
    ).scalars().first()

    if report is None:
        raise NotFoundException("Không tìm thấy báo cáo bảo trì")

    return report


def _build_summary(details: List[BaoCaoBaoTriChiTiet]) -> Dict[str, int]:
    return {
        "tong_yeu_cau": len(details),
        "yeu_cau_da_giai_quyet": sum(
            1 for item in details if item.trang_thai == SuCoBaoTriStatus.HOAN_THANH.value
        ),
    }


def _detail_response(report: BaoCaoBaoTri) -> Dict[str, Any]:
    return {
        "bao_cao": report,
        "chi_tiet": report.chi_tiets,
        "thong_ke": _build_summary(report.chi_tiets),
    }


def create_maintenance_report(
    db: Session,
    payload: BaoCaoBaoTriCreate,
    current_user: Any,
) -> Dict[str, Any]:
    """Lập báo cáo bảo trì theo tháng/năm, tự tổng hợp từ SK_BAOTRI."""
    ma_nhan_vien_lap = _user_attr(current_user, "ma_nv", "ma_nhan_vien")
    if not ma_nhan_vien_lap:
        raise BadRequestException("Không xác định được nhân viên lập báo cáo")

    existed = db.execute(
        select(BaoCaoBaoTri).where(
            BaoCaoBaoTri.thang == payload.thang,
            BaoCaoBaoTri.nam == payload.nam,
        )
    ).scalars().first()
    if existed is not None:
        raise ConflictException(f"Báo cáo bảo trì kỳ {_ky_chot(payload.thang, payload.nam)} đã tồn tại")

    incidents = db.execute(
        select(SuCoBaoTri)
        .where(
            func.month(SuCoBaoTri.ngay_gui) == payload.thang,
            func.year(SuCoBaoTri.ngay_gui) == payload.nam,
        )
        .order_by(SuCoBaoTri.ngay_gui.asc(), SuCoBaoTri.ma_su_co.asc())
    ).scalars().all()

    if not incidents:
        raise BadRequestException(f"Không có dữ liệu sự cố bảo trì cho kỳ {_ky_chot(payload.thang, payload.nam)}")

    ky = _ky_chot(payload.thang, payload.nam)
    report = BaoCaoBaoTri(
        ma_bao_cao=_generate_code("BCBT"),
        ma_nhan_vien_lap=ma_nhan_vien_lap,
        thang=payload.thang,
        nam=payload.nam,
        ky_chot=ky,
        ngay_lap=current_datetime(),
    )

    for index, incident in enumerate(incidents, start=1):
        report.chi_tiets.append(
            BaoCaoBaoTriChiTiet(
                ma_bao_cao=report.ma_bao_cao,
                stt=index,
                ma_yeu_cau=incident.ma_su_co,
                ma_mat_bang=incident.ma_mat_bang,
                ngay_yeu_cau=incident.ngay_gui,
                mo_ta=incident.mo_ta,
                trang_thai=incident.trang_thai,
                ngay_giai_quyet=incident.ngay_hoan_thanh,
                ket_qua=incident.ket_qua,
            )
        )

    actor_ma_tk = _user_attr(current_user, "ma_tk", "ma_tai_khoan")
    with transaction_context(db):
        db.add(report)
        db.flush()
        write_audit_log(
            db=db,
            ma_tk=actor_ma_tk,
            hanh_dong=AuditAction.TAO_MOI,
            doi_tuong="BAOCAOBAOTRI",
            ma_doi_tuong=report.ma_bao_cao,
            chi_tiet=f"Lập báo cáo bảo trì kỳ {ky}",
        )

    report = _get_report_or_404(db, report.ma_bao_cao)
    return _detail_response(report)


def list_maintenance_reports(
    db: Session,
    filters: BaoCaoBaoTriFilter,
    current_user: Optional[Any] = None,
) -> Dict[str, Any]:
    page, page_size = normalize_pagination(filters.page, filters.page_size)
    conditions: List[Any] = []

    if filters.thang is not None:
        conditions.append(BaoCaoBaoTri.thang == filters.thang)
    if filters.nam is not None:
        conditions.append(BaoCaoBaoTri.nam == filters.nam)

    stmt = select(BaoCaoBaoTri)
    count_stmt = select(func.count()).select_from(BaoCaoBaoTri)

    if conditions:
        clause = and_(*conditions)
        stmt = stmt.where(clause)
        count_stmt = count_stmt.where(clause)

    total = db.execute(count_stmt).scalar_one()
    items = db.execute(
        stmt.order_by(BaoCaoBaoTri.nam.desc(), BaoCaoBaoTri.thang.desc(), BaoCaoBaoTri.ngay_lap.desc())
        .offset(calculate_offset(page, page_size))
        .limit(page_size)
    ).scalars().all()

    return {
        "items": items,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": calculate_total_pages(total, page_size),
        },
    }


def get_maintenance_report_detail(
    db: Session,
    ma_bc: str,
    current_user: Optional[Any] = None,
) -> Dict[str, Any]:
    report = _get_report_or_404(db, ma_bc)
    return _detail_response(report)


def export_maintenance_report_excel(
    db: Session,
    ma_bc: str,
    current_user: Optional[Any] = None,
) -> BytesIO:
    report = _get_report_or_404(db, ma_bc)
    summary = _build_summary(report.chi_tiets)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao bao tri"

    headers = [
        "MAYC",
        "MAMB",
        "NGAY YC",
        "MO TA",
        "TRANG THAI",
        "NGAY GIAI QUYET",
        "KET QUA",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1).value = "BÁO CÁO BẢO TRÌ"
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(2, 1).value = f"KỲ CHỐT: {report.ky_chot} | MABC: {report.ma_bao_cao} | MANV: {report.ma_nhan_vien_lap}"
    ws.cell(2, 1).alignment = Alignment(horizontal="center")

    header_row = 4
    fill = PatternFill("solid", fgColor="E2F0D9")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = fill

    for row_index, item in enumerate(report.chi_tiets, start=header_row + 1):
        values = [
            item.ma_yeu_cau,
            item.ma_mat_bang,
            item.ngay_yeu_cau,
            item.mo_ta,
            item.trang_thai,
            item.ngay_giai_quyet,
            item.ket_qua,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_index, col)
            cell.value = value
            if col in (3, 6):
                cell.number_format = "dd/mm/yyyy hh:mm"
            if col == 4:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    stat_start = header_row + len(report.chi_tiets) + 3
    ws.cell(stat_start, 1).value = "THONG KE"
    ws.cell(stat_start, 1).font = Font(bold=True)
    ws.cell(stat_start + 1, 1).value = "TONG YEU CAU"
    ws.cell(stat_start + 1, 2).value = summary["tong_yeu_cau"]
    ws.cell(stat_start + 2, 1).value = "YEU CAU DA GIAI QUYET"
    ws.cell(stat_start + 2, 2).value = summary["yeu_cau_da_giai_quyet"]

    widths = [16, 14, 20, 45, 18, 22, 35]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A5"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
